from openpyxl import Workbook


# a= 10
# def doA():
#     pass

# doA()
# a(10)

wb = Workbook()
print(wb.sheetnames)

m1 = wb.create_sheet()
print(wb.sheetnames)

m1.title = "My시트"
m1.sheet_properties.tabColor='ff66ff'

ws1 = wb.create_sheet()
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
print(wb.sheetnames)

for i in range(1,10):
    wb['Sheet1'][f"A{i}"] = "test"

target = wb.copy_worksheet(wb["Sheet1"])
target.title = "copy"

wb.save("sameple.xlsx")
wb.close()