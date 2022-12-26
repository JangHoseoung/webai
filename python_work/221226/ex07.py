from openpyxl import Workbook #pip install openpyxl
from random import randint

wb = Workbook()
ws = wb.active

ws.append(['번호','국어','영어','수학'])

for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100),randint(0,100)])

print(ws.max_column) #4개 카테고리
print()
print(ws.max_row) #1행부터 11행

a = 2
for row in ws.iter_rows(min_row=2): #2행부터 시작
    total = 0
    for inx,cell in enumerate(row):
        print(inx,cell.value, end = " ")
        if inx != 0:
            total +=int(cell.value)
    ws[f'E{a}'] = total
    a+=1
    print()

ws['E1'] = '총점'

wb.save('test.xlsx')
wb.close()