from openpyxl import Workbook,load_workbook

wb = Workbook()
ws = wb.active

for x in range(1,5):
        Star = '*'
        ws.cell(column=x).value = Star

wb.save("star.xlsx")
wb.close()

lb = load_workbook("star.xlsx")
ws = lb.active 
for x in range(1,10):
        print(ws.cell(column=x).value,end=" ")

print()
lb.close()

