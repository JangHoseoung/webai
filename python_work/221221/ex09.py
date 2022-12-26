from openpyxl import Workbook,load_workbook
import random

# Workbook 파일
# load_workbook 파일 열기
# random 랜덤 값 가져오기
'''
    A10 ~ J10
    1행1열 10행10열까지 랜덤한 값을 넣어서
    random.xlsx로 저장하고
    random.xlsx 불러와서
    1행열 ~ 10행10열까지 출력하기
star.xlsx 파일 저장
...*
...*.*
...*.*.*
...*.*.*.*
...*.*.*.*.*
star.xlsx 파일 불러오기
...*
...*.*
...*.*.*
...*.*.*.*
...*.*.*.*.*
클래스명 star로 작성
1. 메서드 makeStar()
2. 메서드 load star()
'''
wb = Workbook()
ws = wb.active

for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=y,column=x).value = random.randint(0,100)

wb.save("random.xlsx")
wb.close()

lb = load_workbook("random.xlsx")
ws = lb.active
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=y,column=x).value,end=" ")

    print()
lb.close()