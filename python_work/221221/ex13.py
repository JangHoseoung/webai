from openpyxl import Workbook,load_workbook

class Star:
    def makestar(self):
        for i in range(1,6):
            a = "*"
            b = a*i
            print(b)
    
    def doSave(self):
        wb = Workbook()
        ws = wb.active
        for x in range(1,6):
            for y in range(1,16):
                ws.cell(row=y,column=x).value,"*"

        wb.save("star.xlsx")
        wb.close()

    
    def  loadstar(self):
        lb = load_workbook("star.xlsx")
        ws = lb.active
        for x in range(1,6):
            for y in range(1,6):
                print(ws.cell(row=y,column=x).value str="*")

            print()
            lb.close()

        